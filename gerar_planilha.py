#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
gerar_planilha.py — o diário do descomissionamento em Excel.

A planilha NÃO é preenchida à mão. Ela é regenerada do registro: cada linha sai
de uma ata capturada, e cada mudança sai da comparação entre duas atas
consecutivas. Rodar de novo depois de cada ciclo semanal reescreve o arquivo
inteiro a partir da verdade em disco.

Abas:
  Série        uma linha por captura, desde o dia 0
  Mudanças     uma linha por indicador que mudou, com de/para e entre quais datas
  Atestação    quais capturas têm certificado, em que lote, com que hash
  Leia-me      como isto é gerado e o que a atestação prova e não prova

Uso:  python gerar_planilha.py
Saída: DIARIO_DESCOMISSIONAMENTO.xlsx
"""

import hashlib
import json
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
except Exception:
    pass

RAIZ = Path(__file__).resolve().parent
HISTORICO = RAIZ / "carimbo" / "historico"
ATESTACOES = RAIZ / "carimbo" / "atestacoes"
SAIDA = RAIZ / "DIARIO_DESCOMISSIONAMENTO.xlsx"

sys.path.insert(0, str(RAIZ / "site"))
from edicao import indicadores  # noqa: E402  (parser DSR já validado)

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

VERDE = "0B5C4F"
CINZA = "F2F2EF"
AMBAR = "F0B53D"
BRANCO = "FFFFFF"

CAMPOS = [
    ("total_offshore", "Total offshore"),
    ("ciclo.operando", "Operando"),
    ("ciclo.extensao_vida", "Extensão de vida útil"),
    ("ciclo.aguarda_descom", "Aguarda descomissionamento"),
    ("ciclo.em_descom", "Em descomissionamento"),
    ("ciclo.descomissionada", "Descomissionada"),
    ("ciclo.hibernacao", "Hibernação"),
    ("ciclo.em_construcao", "Em construção"),
    ("ciclo.outras", "Outras situações"),
    ("pdi_aprovados", "PDI aprovados"),
    ("pdi_recebidos", "PDI recebidos"),
    ("rdi_aprovados", "RDI aprovados"),
]


def valor(ind, chave):
    if chave.startswith("ciclo."):
        return ind["ciclo"][chave.split(".", 1)[1]]
    return ind[chave]


def sha256_hex(p: Path) -> str:
    h = hashlib.sha256()
    with open(p, "rb") as f:
        for b in iter(lambda: f.read(1 << 20), b""):
            h.update(b)
    return h.hexdigest()


def cabecalho(ws, titulos, largura):
    ws.append(titulos)
    for i, t in enumerate(titulos, 1):
        c = ws.cell(row=1, column=i)
        c.font = Font(bold=True, color=BRANCO, size=10)
        c.fill = PatternFill("solid", fgColor=VERDE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = largura[i - 1]
    ws.row_dimensions[1].height = 34
    ws.freeze_panes = "B2"


def main() -> int:
    if not HISTORICO.is_dir():
        print("histórico não existe"); return 1
    snaps = sorted(d for d in HISTORICO.iterdir()
                   if d.is_dir() and (d / "ata_snapshot.json").is_file())
    if not snaps:
        print("nenhuma captura"); return 1

    # ---- lê cada captura e monta a série ------------------------------------
    serie = []
    for d in snaps:
        try:
            ind = indicadores(d)
        except Exception as e:
            print(f"  ignorada (parse falhou): {d.name} — {e}")
            continue
        cert = d / "ATESTADOS" / "ata_snapshot-autocontido.json"
        assinado = False
        if cert.is_file():
            c = json.loads(cert.read_text(encoding="utf-8"))
            assinado = bool(c.get("issuer_signature"))
        lote = ""
        if ATESTACOES.is_dir():
            for pasta in sorted(ATESTACOES.iterdir()):
                if (pasta / f"{d.name}-autocontido.json").is_file():
                    lote = pasta.name.replace("semana_", "")
                    break
        ind["_pasta"] = d.name
        ind["_atestado"] = assinado
        ind["_lote"] = lote
        ind["_ata_sha"] = sha256_hex(d / "ata_snapshot.json")
        serie.append(ind)

    if not serie:
        print("nenhuma captura legível"); return 1

    dia0 = datetime.fromisoformat(serie[0]["captura_utc"].replace("Z", "+00:00")).date()

    wb = Workbook()

    # ---- aba 1: Série -------------------------------------------------------
    ws = wb.active
    ws.title = "Série"
    titulos = (["Data (UTC)", "Dia", "Captura", "Atestada", "Lote"]
               + [rot for _, rot in CAMPOS]
               + ["Última aprovação PDI", "Última aprovação RDI", "SHA-256 da ata"])
    larg = [12, 6, 24, 10, 12] + [13] * len(CAMPOS) + [18, 18, 26]
    cabecalho(ws, titulos, larg)

    borda = Border(bottom=Side(style="thin", color="DDDDDD"))
    anterior = None
    for i, ind in enumerate(serie):
        dt = datetime.fromisoformat(ind["captura_utc"].replace("Z", "+00:00"))
        linha = [dt.strftime("%d/%m/%Y"), (dt.date() - dia0).days, ind["_pasta"],
                 "sim" if ind["_atestado"] else "não", ind["_lote"]]
        linha += [valor(ind, ch) for ch, _ in CAMPOS]
        linha += [ind["ultima_aprovacao_pdi"], ind["ultima_aprovacao_rdi"],
                  ind["_ata_sha"][:16] + "…"]
        ws.append(linha)
        r = ws.max_row
        for col in range(1, len(titulos) + 1):
            ws.cell(row=r, column=col).border = borda
            ws.cell(row=r, column=col).alignment = Alignment(horizontal="center")
        # destaca a célula que mudou em relação à captura anterior
        if anterior:
            for j, (ch, _) in enumerate(CAMPOS):
                if valor(ind, ch) != valor(anterior, ch):
                    ws.cell(row=r, column=6 + j).fill = PatternFill("solid", fgColor=AMBAR)
                    ws.cell(row=r, column=6 + j).font = Font(bold=True)
        if i % 2 == 1:
            for col in range(1, 6):
                if not ws.cell(row=r, column=col).fill.fgColor.rgb == "00" + AMBAR:
                    ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=CINZA)
        anterior = ind

    # ---- aba 2: Mudanças ----------------------------------------------------
    ws2 = wb.create_sheet("Mudanças")
    cabecalho(ws2, ["De (data)", "Para (data)", "Dias", "Indicador", "De", "Para",
                    "Variação", "Captura de origem", "Captura de destino",
                    "As duas atestadas"],
              [12, 12, 7, 30, 9, 9, 10, 24, 24, 16])
    mudancas = 0
    for a, b in zip(serie, serie[1:]):
        da = datetime.fromisoformat(a["captura_utc"].replace("Z", "+00:00")).date()
        db = datetime.fromisoformat(b["captura_utc"].replace("Z", "+00:00")).date()
        for ch, rot in CAMPOS:
            va, vb = valor(a, ch), valor(b, ch)
            if va != vb:
                ws2.append([da.strftime("%d/%m/%Y"), db.strftime("%d/%m/%Y"),
                            (db - da).days, rot, va, vb, vb - va,
                            a["_pasta"], b["_pasta"],
                            "sim" if (a["_atestado"] and b["_atestado"]) else "não"])
                r = ws2.max_row
                for col in range(1, 11):
                    ws2.cell(row=r, column=col).alignment = Alignment(horizontal="center")
                ws2.cell(row=r, column=7).font = Font(bold=True)
                mudancas += 1
        for campo, rot in (("ultima_aprovacao_pdi", "Data da última aprovação de PDI"),
                           ("ultima_aprovacao_rdi", "Data da última aprovação de RDI")):
            if a[campo] != b[campo]:
                ws2.append([da.strftime("%d/%m/%Y"), db.strftime("%d/%m/%Y"),
                            (db - da).days, rot, a[campo], b[campo], "",
                            a["_pasta"], b["_pasta"],
                            "sim" if (a["_atestado"] and b["_atestado"]) else "não"])
                mudancas += 1
    if mudancas == 0:
        ws2.append(["", "", "", "Nenhuma mudança registrada até agora", "", "", "", "", "", ""])

    # ---- aba 3: Eventos (nível de registro, que é a verdade) ----------------
    # O agregado diz que um número mexeu; só o diff por registro diz o QUE mexeu.
    # Sem isto, sonda trocando de poço vira "tendência".
    sys.path.insert(0, str(RAIZ))
    from diff_capturas import comparar  # noqa: E402

    wsE = wb.create_sheet("Eventos")
    cabecalho(wsE, ["De (data)", "Para (data)", "Tipo de evento", "Registro",
                    "De", "Para", "Relevante"],
              [12, 12, 34, 30, 34, 34, 12])
    eventos = 0
    for a, b in zip(serie, serie[1:]):
        da = datetime.fromisoformat(a["captura_utc"].replace("Z", "+00:00")).date()
        db = datetime.fromisoformat(b["captura_utc"].replace("Z", "+00:00")).date()
        for tipo, chave, de, para in comparar(HISTORICO / a["_pasta"],
                                              HISTORICO / b["_pasta"], verboso=False):
            # sonda mudando de poço não é evento de ciclo de vida: ver Leia-me
            operacional = any(str(x).startswith(("Perfuração", "Abandono ", "Completação",
                                                 "Intervenção", "Avaliação", "Workover"))
                              for x in (de, para))
            wsE.append([da.strftime("%d/%m/%Y"), db.strftime("%d/%m/%Y"), tipo,
                        chave, str(de), str(para),
                        "não, é operação de poço" if operacional else "SIM"])
            r = wsE.max_row
            for col in range(1, 8):
                wsE.cell(row=r, column=col).alignment = Alignment(
                    horizontal="center", wrap_text=True)
            if not operacional:
                wsE.cell(row=r, column=7).fill = PatternFill("solid", fgColor=AMBAR)
                wsE.cell(row=r, column=7).font = Font(bold=True)
            eventos += 1
    if eventos == 0:
        wsE.append(["", "", "Nenhum evento por registro até agora", "", "", "", ""])

    # ---- aba 4: Atestação ---------------------------------------------------
    ws3 = wb.create_sheet("Atestação")
    cabecalho(ws3, ["Captura", "Data (UTC)", "Lote", "Certificado assinado",
                    "SHA-256 da ata", "Elo da cadeia"],
              [24, 12, 14, 20, 68, 26])
    for ind in serie:
        ata = json.loads((HISTORICO / ind["_pasta"] / "ata_snapshot.json")
                         .read_text(encoding="utf-8"))
        cad = ata.get("cadeia")
        elo = cad["snapshot_anterior"] if cad else "gênese da série"
        dt = datetime.fromisoformat(ind["captura_utc"].replace("Z", "+00:00"))
        ws3.append([ind["_pasta"], dt.strftime("%d/%m/%Y"), ind["_lote"],
                    "sim" if ind["_atestado"] else "não", ind["_ata_sha"], elo])
        for col in range(1, 7):
            ws3.cell(row=ws3.max_row, column=col).alignment = Alignment(horizontal="center")

    # ---- aba 4: Leia-me -----------------------------------------------------
    ws4 = wb.create_sheet("Leia-me")
    ws4.column_dimensions["A"].width = 118
    texto = [
        ("Diário do Descomissionamento, série da ANP", True),
        ("", False),
        (f"Gerado em {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M UTC')} "
         f"por gerar_planilha.py, a partir de {len(serie)} capturas em disco.", False),
        (f"Dia 0 da série: {dia0.strftime('%d/%m/%Y')}.", False),
        ("", False),
        ("NÃO EDITE ESTA PLANILHA À MÃO.", True),
        ("Ela é reescrita inteira a cada execução, a partir das atas capturadas. "
         "Qualquer valor digitado aqui some na próxima geração, e pior, deixa de "
         "ter lastro. Se um número precisa mudar, o que muda é a fonte.", False),
        ("", False),
        ("Como cada linha nasce", True),
        ("Todo dia, às 11h30, uma rotina consulta o painel público de "
         "descomissionamento da ANP e grava a resposta bruta do servidor. Cada "
         "resposta é fixada por SHA-256 numa ata datada, e as atas mais novas "
         "registram também o hash da anterior, encadeando a série. Os indicadores "
         "desta planilha são recalculados dessas respostas, não digitados.", False),
        ("", False),
        ("O que a atestação prova", True),
        ("Uma vez por semana as atas vão para hardware seguro (TEE), que devolve "
         "um certificado. Ele prova três coisas: que aquela ata é exatamente "
         "aquela, que passou por hardware genuíno rodando um programa cujo código "
         "é medido, e que quem confere não precisa confiar em mim. A verificação "
         "roda sem internet, contra a raiz de confiança do fabricante do "
         "processador.", False),
        ("", False),
        ("O que a atestação NÃO prova", True),
        ("Não prova de onde o conteúdo veio: a origem fica declarada na própria "
         "ata, com a consulta escrita ali para qualquer pessoa refazer. Não "
         "carimba data: o certificado não tem campo de tempo, e isso é "
         "deliberado. E não avalia frescor, isto é, se a plataforma foi revogada "
         "ou se o nível de segurança dela estava em dia no momento da emissão. O "
         "que está provado é a autenticidade, que é permanente; o frescor é uma "
         "afirmação com data, e exige congelar no certificado o material que o "
         "fabricante publica. Isso está identificado e é implementável.", False),
        ("", False),
        ("Cuidado ao ler a linha 'Operando'", True),
        ("O campo de situação do cadastro não guarda só estado de ciclo de vida. "
         "Para unidade móvel, ele guarda a operação em curso, com o nome do poço: "
         "'Perfuração 8-SPH-27D-SPS', 'Abandono 7-BR-49HPA-RJS', 'Completação "
         "7-JUB-85H-ESS'. Dos valores distintos de situação vistos no cadastro "
         "offshore, a maior parte é desse tipo. Consequência prática: a contagem "
         "de 'Operando' oscila de um dia para o outro sem que nada tenha mudado "
         "no ciclo de vida da unidade, e por isso essa linha não serve de "
         "indicador. As linhas que interessam a este diário, aguarda "
         "descomissionamento, em descomissionamento e descomissionada, descrevem "
         "estado terminal de plataforma fixa e não sofrem esse efeito.", False),
        ("", False),
        ("Fonte e base pública", True),
        ("Painel Dinâmico de Descomissionamento de Instalações de Exploração e "
         "Produção da ANP. A base tratada do estudo que originou este diário está "
         "publicada em doi.org/10.5281/zenodo.21919402, sob licença aberta.", False),
    ]
    for t, negrito in texto:
        ws4.append([t])
        c = ws4.cell(row=ws4.max_row, column=1)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if negrito:
            c.font = Font(bold=True, size=11, color=VERDE)

    wb.save(SAIDA)
    atest = sum(1 for i in serie if i["_atestado"])
    print(f"capturas na série : {len(serie)}  (dia 0 = {dia0.strftime('%d/%m/%Y')})")
    print(f"atestadas         : {atest}")
    print(f"mudanças agregadas: {mudancas}")
    print(f"eventos por registro: {eventos}")
    print(f"gerado            : {SAIDA.name}  ({SAIDA.stat().st_size:,} bytes)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
